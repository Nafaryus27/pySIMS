import openpyxl
import numpy as np
import matplotlib.pyplot as plt

path = "/home/garcia/Stage/isotope/"
filename = "tableau.xlsx"

def create_dico(filename):
    
    book = openpyxl.load_workbook(filename)
    sheet = book.active

    datadico = {}

    for ligne in range(11,133):

        # Ignore les lignes vides
        if sheet.cell(ligne,1).value == None:
            continue

        try:            
            

            raw_data = []
            isotope = True
            
            current_cell = sheet.cell(ligne,1).value    # Z, melting
            current_cell = current_cell.split()
            for x in current_cell :
                raw_data.append(x)
                

            current_cell = sheet.cell(ligne,2).value    # nom, lettre, masse_atomique
            current_cell = current_cell.split()
            for x in current_cell :
                raw_data.append(x)


            current_cell = sheet.cell(ligne,3).value    # ionisation, affinite
            current_cell = current_cell.split()
            for x in current_cell :
                raw_data.append(x)


            current_cell = sheet.cell(ligne,4).value    # masse_entiere (plusieurs)
            try :    # Si un seul isotope
                raw_data.append(int(current_cell))
            except :
                if current_cell == 'no naturally occurring isotopes' :
                    isotope = False
                    
                if isotope :
                    current_cell = current_cell.split()
                    for x in current_cell :
                        raw_data.append(x)

            
            current_cell = sheet.cell(ligne,5).value    # masse_isotope (plusieurs)
            try :    # Si un seul isotope
                raw_data.append(int(current_cell))
            except :
                if isotope :
                    current_cell = current_cell.split()
                    for x in current_cell :
                        raw_data.append(x)


            current_cell = sheet.cell(ligne,6).value    # abondance (plusieurs)
            try :    # Si un seul isotope
                raw_data.append(int(current_cell))
            except :
                if isotope :
                    current_cell = current_cell.split()
                    for x in current_cell :
                        raw_data.append(x)


            current_cell = sheet.cell(ligne,7).value    # source
            if isotope :
                current_cell = current_cell.split()
                nbr_sources = len(current_cell)
                for x in current_cell :
                    raw_data.append(x)
            else :
                nbr_sources = 0



            for i in range(len(raw_data)) :
                try :
                    raw_data[i] = float(raw_data[i])
                    if raw_data[i] -int(raw_data[i]) == 0.0 :
                        raw_data[i] = int(raw_data[i])
                except :
                    pass

            #   ----- POST TRAITEMENT -----

            # Comptage du nombre d'isotopes
            if isotope :
                nbr_isotopes = (len(raw_data) - nbr_sources - 7)//3
            else :
                nbr_isotopes = 0

            dico_element = {'Z' : raw_data[0],
                            'melting' : raw_data[1],
                            'nom' : raw_data[2],
                            'lettre' : raw_data[3],
                            'masse_atomique' : raw_data[4],
                            'ionisation' : raw_data[5],
                            'affinite' : raw_data[6]}


            # Isotopes
            if nbr_isotopes > 0 :
                dico_element['isotope'] = {}
            for i in range(nbr_isotopes):
                dico_element['isotope'][i+1] = {'masse_entiere' : raw_data[7+i],
                                                     'masse' : raw_data[7+nbr_isotopes+i],
                                                     'abondance' : raw_data[7+(2*nbr_isotopes)+i]}

            # Sources
            if nbr_sources > 0 :
                dico_element['source'] = raw_data[-nbr_sources:]

            datadico[dico_element['lettre']] = dico_element

        except :
           f"Erreur ligne {ligne}"

    return datadico





'''
print(Elements['La'])
'''
###################################################################################################

def get_isotope(Elements,elem):

  
    n = len(Elements[elem]['isotope'])
    
    x = [ Elements[elem]['isotope'][i+1]['masse'] for i in range(n) ]
    
    
    y = [ Elements[elem]['isotope'][i+1]['abondance'] for i in range(n) ]
    
    
    z = [ (str(round(x[i]))+elem) for i in range(len(x)) ]
    

    
    valeur = {}
    
    for i in range(len(x)):
        valeur[z[i]]=[]
    
    for i in range(len(x)):
        valeur[z[i]].append(x[i])
        valeur[z[i]].append(y[i]*0.01)
        
#    for i in valeur:
#       print(i,":",valeur[i])


    return valeur
  
##################################################################################################
#def graph_isotope(Elements,elem):
#
#
#    n = len(Elements[elem]['isotope'])
#    
#    x = [ Elements[elem]['isotope'][i+1]['masse'] for i in range(n) ]
#        
#    y = [ Elements[elem]['isotope'][i+1]['abondance'] for i in range(n) ]
#
#       
#    for i in range(len(x)) :
#        plt.plot( [x[i],x[i]], [0,y[i]], color='darkgreen')
#        plt.text(x[i]-0.02, 0.01, r"$^{"+str(round(x[i]))+"}$"+elem)   
#        plt.xlabel("Masse")
#        plt.ylabel("Abondance (%)")
#        


##################################################################################################
def get_complexe(Elements,elem1,elem2):


    masse1 = [ Elements[elem1]['isotope'][i+1]['masse'] for i in range(len(Elements[elem1]['isotope'])) ]
    #print(x1)
    masse2 = [ Elements[elem2]['isotope'][i+1]['masse'] for i in range(len(Elements[elem2]['isotope'])) ] 
    #print(x2)
     
    abon1 = [ Elements[elem1]['isotope'][i+1]['abondance'] for i in range(len(Elements[elem1]['isotope'])) ]
    abon1 = [i*0.01 for i in abon1]
    #print(y1)
    
    abon2 = [ Elements[elem2]['isotope'][i+1]['abondance'] for i in range(len(Elements[elem2]['isotope'])) ] 
    abon2 = [i*0.01 for i in abon2]
    #print(y2)
    
    nom1 = [ (str(round(masse1[i]))+elem1) for i in range(len(masse1)) ]
    #print(z1)
    nom2 = [ (str(round(masse2[i]))+elem2) for i in range(len(masse2)) ]
    #print(z1)
    
    masse = []
    abon =  []
    couple = []
    
    if elem1 == elem2:
        for i in range(len(masse1)):
            for j in range(i, len(masse2)):
                masse.append(masse1[i]+masse2[j])
                abon.append(abon1[i]*abon2[j])
                couple.append(nom1[i]+nom2[j])        
    else:
        for i in range(len(masse1)):
            for j in range( len(masse2)):
                masse.append(masse1[i]+masse2[j])
                abon.append(abon1[i]*abon2[j])
                couple.append(nom1[i]+nom2[j])


    binaire = {}
    
    for i in range(len(masse)):
        binaire[couple[i]]=[]
    
    for i in range(len(masse)):
        binaire[couple[i]].append(masse[i])
        binaire[couple[i]].append(abon[i])
       
#    for k,v in sorted(binaire.items(), key = lambda  binaire: binaire[1],reverse=True):
#        print("%s: %s"% (k,v))

    return binaire
##################################################################################################
#def graph_complexe(Elements,elem1,elem2):
#
#
#    m1 = [ Elements[elem1]['isotope'][i+1]['masse'] for i in range(len(Elements[elem1]['isotope'])) ]
#    #print(x1)
#    m2 = [ Elements[elem2]['isotope'][i+1]['masse'] for i in range(len(Elements[elem2]['isotope'])) ] 
#    #print(x2)
#     
#    a1 = [ Elements[elem1]['isotope'][i+1]['abondance'] for i in range(len(Elements[elem1]['isotope'])) ]
#    a1 = [i*0.01 for i in a1]
#    #print(y1)
#    
#    a2 = [ Elements[elem2]['isotope'][i+1]['abondance'] for i in range(len(Elements[elem2]['isotope'])) ] 
#    a2 = [i*0.01 for i in a2]
#    #print(y2)139La18O: [156.905506, 0.0019982]
#    
#    
#    masse = []
#    abon =  []
#    
#    for i in range(len(m1)):
#        for j in range(len(m2)):
#            masse.append(m1[i]+m2[j])
#    
#    for i in range(len(a1)):
#        for j in range(len(a2)):
#            abon.append(a1[i]*a2[j])
#            
#    for i in range(len(masse)) :
#        plt.semilogy([masse[i],masse[i]], [0,abon[i]])
#    
#    
#    plt.xlabel("Masse")
#    plt.ylabel("Abondance (%)")
    
##################################################################################################
    
def merge_dico(dico1,dico2):
    final_dico = dict()
    for dictionnaire in [dico1,dico2]:
        for key,value in dictionnaire.items():
            final_dico[key] = value
    
    return final_dico

##################################################################################################
def print_spectrum(Elements,elems):

    list_isotope = dict()
    
    
    for i in range(len(elems)):
        list_isotope = merge_dico(list_isotope, get_isotope(Elements,elems[i]))
       
    
    
    list_complexe_duo = dict()
    
    for i in range(0,len(elems)):
        for j in range (i, len(elems)): 
            list_complexe_duo = merge_dico(list_complexe_duo, get_complexe(Elements,elems[i],elems[j]))
       
    full_list = merge_dico(list_isotope, list_complexe_duo)    
    
    for k,v in sorted(full_list.items(), key = lambda  full_list: full_list[1],reverse=True):#[1] our classer par masse                   
        print("%s: %s"% (k,v))      
    
    return None                                                           #[1][1] pour classer par abondance


##################################################################################################
def extract_spectrum_mono(Elements,elems):

    list_isotope = dict()    
    
    for i in range(len(elems)):
        list_isotope = merge_dico(list_isotope, get_isotope(Elements,elems[i]))
          
    return list_isotope

def extract_spectrum_complex(Elements,elems):

    list_complexe = dict()
    
    for i in range(0,len(elems)):
        for j in range (i, len(elems)): 
            list_complexe = merge_dico(list_complexe, get_complexe(Elements,elems[i],elems[j]))
    
    return list_complexe
##################################################################################################
def convert_spectrum_dict_iso_into_tuple (full_dict_iso) :
    # convertit le dictionnaire 'complex' : [mass, abondance] en tuples
    # pour affichage scatter
    isotop_c = ()
    masses_iso = ()
    abondances_iso = ()
    for isotop in full_dict_iso.keys () :
        isotop_c += isotop,
        masses_iso += full_dict_iso [isotop][0],
        abondances_iso += full_dict_iso [isotop][1], 
    return isotop_c, masses_iso, abondances_iso

def convert_spectrum_dict_complex_into_tuple (full_dict_complex) :

    ionic_c = ()
    masses_complex = ()
    abondances_complex = ()
    for ionic_complex in full_dict_complex.keys () :
        ionic_c += ionic_complex,
        masses_complex += full_dict_complex [ionic_complex][0],
        abondances_complex += full_dict_complex [ionic_complex][1], 
    return ionic_c, masses_complex, abondances_complex



def main () : 
    path = "/home/garcia/Stage/isotope/"
    filename = "tableau.xlsx"
    Elements = create_dico(filename)
    #print_spectrum(Elements,['Ba','La','Ti','Ni','Sr','H','O','C'])
    full_dict_iso = extract_spectrum_mono(Elements,['Ba','La','Ti','Ni','Sr','H','O','C','Cs'])
    isotop_c, masses_iso, abondances_iso = convert_spectrum_dict_iso_into_tuple (full_dict_iso)
    full_dict_complex = extract_spectrum_complex(Elements,['Ba','La','Ti','Ni','Sr','H','O','C','Cs'])
    ionic_c, masses_complex, abondances_complex = convert_spectrum_dict_complex_into_tuple (full_dict_complex)
    fig1 = plt.figure ()
    ax = fig1.add_subplot (111)
    ax.scatter (masses_complex, abondances_complex, marker = '1')

    
    
if __name__ == '__main__' :
    main()
    
##################################################################################################    
'''
   print(Elements['elem'])   
   
   get_isotope('elem') --> Affiche la masse et l'abondance des isotopes de l'element 'elem'
   
   get_complexe('elem1','elem2') --> Affiche la masse et l'abondance des complexes binaires des 
   elements 'elem1' et 'elem2'
   
   graph_isotope('elem') --> Affiche le graphique Abondance =f(masse) des isotopes de 'elem'
   
   graph_complexe('elem1','elem2')  --> Affiche le graphique Abondance =f(masse) des isotopes des
   complexes binaires des éléments 'elem1' et 'elem2'

'''

##################################################################################################
